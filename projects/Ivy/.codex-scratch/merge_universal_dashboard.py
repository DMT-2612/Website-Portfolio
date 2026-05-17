from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


UNIVERSAL = Path(r"C:/Users/trinh/Downloads/Universal.xlsx")
DASHBOARD = Path(r"C:/Users/trinh/Downloads/DASHBOARD FILE.xlsx")
OUTPUT = Path(r"C:/Users/trinh/OneDrive/Desktop/JOB/Ivy/Universal.xlsx")

UNIVERSAL_SHEET = "Sheet1"
DASHBOARD_SHEET = "ORIGNAL FILE"
MERGED_SHEET = "Merged Dataset"

UNIVERSAL_HEADERS = [
    "Date",
    "Order Number",
    "Name",
    "Sale Channels",
    "Product",
    "Animals",
    "Colour",
    "Occasions",
    "Shipping Method",
    "Shipping Fee",
    "Product Price",
    "Total",
    "State",
    "Email",
]


def normalize_header(value):
    return str(value).strip() if value is not None else ""


def year_from_date(value):
    if isinstance(value, datetime):
        return str(value.year)
    text = str(value or "").strip()
    if len(text) >= 4 and text[:4].isdigit():
        return text[:4]
    return "unknown"


def read_rows(path, sheet_name):
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name]
    headers = [normalize_header(cell.value) for cell in ws[1]]
    index = {header: pos for pos, header in enumerate(headers) if header}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        rows.append((index, row))
    return rows


def build_universal_rows():
    rows = []
    for index, row in read_rows(UNIVERSAL, UNIVERSAL_SHEET):
        rows.append([row[index[h]] if h in index and index[h] < len(row) else None for h in UNIVERSAL_HEADERS])
    return rows


def build_dashboard_rows():
    rows = []
    mapping = {
        "Date": "Date",
        "Name": "Name",
        "Sale Channels": "Sale Channels",
        "Product": "Products",
        "Animals": "Animals",
        "Colour": "Colour",
        "Occasions": "Occasions",
        "Shipping Method": "Shipping Method",
        "Shipping Fee": "Shipping Fee",
        "Product Price": "Product Price",
        "Total": "Total",
        "State": "State",
        "Email": "Email",
    }
    for counter, (index, row) in enumerate(read_rows(DASHBOARD, DASHBOARD_SHEET), start=1):
        date_value = row[index["Date"]] if "Date" in index and index["Date"] < len(row) else None
        order_number = f"old-{year_from_date(date_value)}-{counter:03d}"
        merged = []
        for header in UNIVERSAL_HEADERS:
            if header == "Order Number":
                merged.append(order_number)
                continue
            source = mapping[header]
            merged.append(row[index[source]] if source in index and index[source] < len(row) else None)
        rows.append(merged)
    return rows


def copy_header_style(source_ws, target_ws):
    for col_idx, source_cell in enumerate(source_ws[1], start=1):
        target_cell = target_ws.cell(row=1, column=col_idx)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
        target_cell.value = UNIVERSAL_HEADERS[col_idx - 1]


def main():
    universal_rows = build_universal_rows()
    dashboard_rows = build_dashboard_rows()

    wb = load_workbook(UNIVERSAL)
    if MERGED_SHEET in wb.sheetnames:
        del wb[MERGED_SHEET]
    source_ws = wb[UNIVERSAL_SHEET]
    ws = wb.create_sheet(MERGED_SHEET, 1)

    copy_header_style(source_ws, ws)

    for row_idx, values in enumerate(universal_rows + dashboard_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            source_cell = source_ws.cell(row=min(row_idx, source_ws.max_row), column=col_idx)
            if source_cell.has_style:
                cell.number_format = source_cell.number_format

    for col_idx, source_col in enumerate(source_ws.column_dimensions.values(), start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if source_col.width:
            ws.column_dimensions[letter].width = source_col.width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    table_ref = f"A1:N{ws.max_row}"
    table = Table(displayName="MergedDataset", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(OUTPUT)
    print(f"output={OUTPUT}")
    print(f"universal_rows={len(universal_rows)}")
    print(f"dashboard_rows={len(dashboard_rows)}")
    print(f"merged_data_rows={len(universal_rows) + len(dashboard_rows)}")
    print(f"merged_total_rows_with_header={ws.max_row}")


if __name__ == "__main__":
    main()
